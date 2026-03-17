from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any, Literal, Optional

import pandas as pd

from scripts.utils.logger import get_logger


logger = get_logger("file_utils")


# ─── Directory helpers ────────────────────────────────────────────────────────


def ensure_parent_dir(path: str | Path) -> None:
    Path(path).parent.mkdir(parents=True, exist_ok=True)


# ─── Encoding / separator detection ──────────────────────────────────────────


def detect_encoding(path: str | Path) -> str:
    """Detect the encoding of a text file.

    Strategy:
    1. Check for UTF-8 BOM → return "utf-8-sig"
    2. Try strict UTF-8 decode → return "utf-8"
    3. Try CP1252 → return "cp1252"
    4. Fall back to latin-1 (never raises) → return "latin-1"
    """
    raw = Path(path).read_bytes()

    if raw.startswith(b"\xef\xbb\xbf"):
        return "utf-8-sig"

    try:
        raw.decode("utf-8")
        return "utf-8"
    except UnicodeDecodeError:
        pass

    try:
        raw.decode("cp1252")
        return "cp1252"
    except UnicodeDecodeError:
        pass

    return "latin-1"


def detect_separator(path: str | Path, encoding: Optional[str] = None) -> str:
    """Detect the column separator in a CSV-like file.

    Strategy:
    1. Use csv.Sniffer on the first 4 KB.
    2. If Sniffer fails, count occurrences of candidate separators in the
       first 20 lines and pick the most frequent one.

    Candidates: ; , ~ \\t |
    Default fallback: ,
    """
    enc = encoding or detect_encoding(path)
    candidates = [";", ",", "~", "\t", "|"]

    try:
        with open(path, "r", encoding=enc, newline="") as fh:
            sample = fh.read(4096)
        dialect = csv.Sniffer().sniff(sample, delimiters="".join(candidates))
        return dialect.delimiter
    except csv.Error:
        pass

    # Frequency-count fallback
    counts: dict[str, int] = {sep: 0 for sep in candidates}
    with open(path, "r", encoding=enc) as fh:
        for i, line in enumerate(fh):
            if i >= 20:
                break
            for sep in candidates:
                counts[sep] += line.count(sep)

    best = max(counts, key=lambda s: counts[s])
    return best if counts[best] > 0 else ","


# ─── CSV ──────────────────────────────────────────────────────────────────────


def read_csv_file(
    path: str | Path,
    encoding: Optional[str] = None,
    separator: Optional[str] = None,
) -> pd.DataFrame:
    """Read a CSV file with automatic encoding and separator detection."""
    enc = encoding or detect_encoding(path)
    sep = separator or detect_separator(path, encoding=enc)
    logger.debug("Reading CSV %s (encoding=%s, sep=%r)", path, enc, sep)
    return pd.read_csv(path, encoding=enc, sep=sep)


def write_csv_file(df: pd.DataFrame, path: str | Path) -> None:
    ensure_parent_dir(path)
    df.to_csv(path, index=False)


# ─── Excel ────────────────────────────────────────────────────────────────────


def _detect_header_row(sheet: Any, max_scan: int = 20) -> int:
    """Return the 0-based row index where the actual header starts.

    Heuristic: the first row that has more than half its cells populated
    and at least 2 non-empty cells is treated as the header.
    """
    for i in range(min(max_scan, sheet.max_row)):
        row_values = [
            cell.value
            for cell in sheet[i + 1]
            if cell.value is not None and str(cell.value).strip() != ""
        ]
        if len(row_values) >= 2:
            return i
    return 0


def read_excel_file(
    path: str | Path,
    sheet_name: Optional[str | int] = 0,
    header_row: Optional[int] = None,
    auto_detect_header: bool = False,
) -> pd.DataFrame:
    """Read a sheet from an Excel file (.xlsx / .xls).

    Parameters
    ----------
    path:
        Path to the Excel file.
    sheet_name:
        Sheet name or 0-based index. Defaults to the first sheet (0).
    header_row:
        Explicit 0-based row index for the header. Overrides auto-detection.
    auto_detect_header:
        When True, scan the sheet to find the first data row automatically.
        Ignored if *header_row* is provided.
    """
    file_path = Path(path)
    if not file_path.exists():
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    if header_row is None and auto_detect_header:
        import openpyxl  # noqa: PLC0415

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet_key = (
            sheet_name
            if isinstance(sheet_name, str)
            else wb.sheetnames[sheet_name or 0]
        )
        header_row = _detect_header_row(wb[sheet_key])
        wb.close()
        logger.debug("Auto-detected header row %d in sheet %r", header_row, sheet_key)

    effective_header = header_row if header_row is not None else 0
    raw = pd.read_excel(
        file_path,
        sheet_name=sheet_name,
        header=effective_header,
        engine="openpyxl",
    )
    # pd.read_excel returns dict[str, DataFrame] when sheet_name is None;
    # we always pass a concrete sheet so the result is a single DataFrame.
    if isinstance(raw, dict):
        raise ValueError(
            "read_excel_file received multiple sheets. "
            "Use read_excel_all_sheets() instead."
        )
    df: pd.DataFrame = raw
    logger.debug(
        "Read Excel %s | sheet=%r | header_row=%d | shape=%s",
        file_path,
        sheet_name,
        effective_header,
        df.shape,
    )
    return df


def read_excel_all_sheets(
    path: str | Path,
    auto_detect_header: bool = False,
) -> dict[str, pd.DataFrame]:
    """Read all sheets from an Excel file.

    Returns a dict mapping sheet name → DataFrame.
    """
    import openpyxl  # noqa: PLC0415

    file_path = Path(path)
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    wb.close()

    result: dict[str, pd.DataFrame] = {}
    for name in sheet_names:
        result[name] = read_excel_file(
            path, sheet_name=name, auto_detect_header=auto_detect_header
        )
    return result


# ─── Parquet ──────────────────────────────────────────────────────────────────


def read_parquet_file(path: str | Path) -> pd.DataFrame:
    """Read a Parquet file."""
    logger.debug("Reading Parquet %s", path)
    return pd.read_parquet(path)


ParquetCompression = Literal["snappy", "gzip", "brotli", "lz4", "zstd"]


def write_parquet_file(
    df: pd.DataFrame,
    path: str | Path,
    compression: Optional[ParquetCompression] = "snappy",
) -> None:
    """Write a DataFrame to Parquet format.

    Parameters
    ----------
    df:
        DataFrame to persist.
    path:
        Destination file path (should end in .parquet).
    compression:
        Parquet compression codec. Options: snappy, gzip, brotli, lz4, zstd.
        Pass None to disable compression. Defaults to snappy.
    """
    ensure_parent_dir(path)
    df.to_parquet(path, index=False, compression=compression)
    logger.debug(
        "Wrote Parquet %s (compression=%s, rows=%d)", path, compression, len(df)
    )


# ─── JSON ─────────────────────────────────────────────────────────────────────


def read_json_file(path: str | Path) -> Any:
    with open(path, "r", encoding="utf-8") as file:
        return json.load(file)


def write_json_file(data: Any, path: str | Path) -> None:
    ensure_parent_dir(path)
    with open(path, "w", encoding="utf-8") as file:
        json.dump(data, file, indent=2, ensure_ascii=False)
